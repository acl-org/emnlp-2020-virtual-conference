import csv
import random
from dataclasses import dataclass
from typing import List, Dict

import pytz
import ruamel
from openpyxl import load_workbook
from ruamel import yaml

from datetime import datetime

import pandas as pd

# https://docs.google.com/spreadsheets/d/1IDk3K1JD1hvH_hvyMy6TeRuE2F6DQDfpgwNpTIP9KgI/edit#gid=1276180010
from scripts.dataentry.paths import *


def load_excel():
    wb = load_workbook(PATH_SOCIALS)

    ws = wb.worksheets[0]
    ws.delete_cols(10, 100)
    df = pd.DataFrame(
        ws.values,
        columns=[
            "ID",
            "Event Type",
            "Event",
            "Pre-recorded/live",
            "Platform",
            "Organizers",
            "Contact person",
            "Email address",
            "Channel Name",
        ],
    )
    df = df.dropna(subset=["ID"])
    df = df[:-1]

    id_to_organizers = {
        row["ID"]: [e.strip() for e in row["Organizers"].split(",")]
        for _, row in df.iterrows()
    }
    id_to_name = {row["ID"]: row["Event"] for _, row in df.iterrows()}
    id_to_channel = {row["ID"]: row["Channel Name"] for _, row in df.iterrows()}
    id_to_location = {row["ID"]: row["Platform"] for _, row in df.iterrows()}

    result = []
    for ws in wb.worksheets[2:]:
        data = {}

        uid = ws["B2"].value
        description = ws["B3"].value
        website = ws["B4"].value
        name = id_to_name[uid]

        print(uid)

        images = {
            "A1": "static/images/socials/queer_in_ai.png",
            "A2": "static/images/socials/VegNLP-logo.png",
            "A3": "static/images/socials/LXAI-navlogo.png",
            "A4": "static/images/socials/NorthAfricansInNLP.png",
        }

        ws.delete_rows(1, 9)
        ws.delete_cols(7, 100)
        df = pd.DataFrame(
            ws.values,
            columns=[
                "Session Name",
                "Day",
                "Start Time",
                "End Time",
                "Time Zone",
                "Host",
            ],
        )
        df.dropna(subset=["Session Name"], inplace=True)

        data["name"] = name
        data["description"] = description
        data["UID"] = uid
        data["organizers"] = {"members": id_to_organizers[uid]}

        if uid in images:
            data["image"] = images[uid]

        if uid.startswith("B"):
            data["image"] = "static/images/emnlp2020/acl-logo.png"

        if website:
            data["website"] = website

        data["rocketchat_channel"] = id_to_channel[uid]
        data["location"] = id_to_location[uid]
        result.append(data)

        sessions = []
        for idx, row in df.iterrows():
            name = "S-" + row["Session Name"].strip()

            if uid.startswith("B") and row["Host"]:
                name = name + " with " + row["Host"]

            day = row["Day"]
            start_time = row["Start Time"]

            if isinstance(start_time, datetime):
                start_time = start_time.time()

            end_time = row["End Time"]
            # assert row["Time Zone"] == "UTC-0", "Was" + str(row["Time Zone"] )

            tz = pytz.utc

            start = datetime.combine(day.date(), start_time)
            start = tz.localize(start)

            end = datetime.combine(day.date(), end_time)
            end = tz.localize(end)

            sessions.append(
                {"name": name, "start_time": start, "end_time": end,}
            )

        data["sessions"] = sessions

    result.sort(key=lambda x: x["UID"])
    yaml.scalarstring.walk_tree(result)

    with open("yamls/socials.yml", "w") as f:
        yaml.dump(result, f, Dumper=ruamel.yaml.RoundTripDumper)


if __name__ == "__main__":
    download_socials()
    load_excel()
