from collections import defaultdict
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any
from typing import List, Dict
import xml.etree.ElementTree as ET

import pandas as pd
import pytz
import ruamel
from ftfy import fix_text
from openpyxl import load_workbook
from pylatexenc.latex2text import LatexNodes2Text
from ruamel import yaml
from ruamel import yaml

# https://docs.google.com/spreadsheets/d/19LRnJpae5NQd0D1NEO40kTbwDvS9f125tpsjBdevrcs/edit#gid=0
from scripts.dataentry.paths import *

fix = {
    "490": "4th Workshop on Structured Prediction for NLP",
    "510": "CoNLL 2020",
    "884": "Deep Learning Inside Out (DeeLIO): The First Workshop on Knowledge Extraction and Integration for Deep Learning Architectures",
    "1093": "SIGTYP 2020: The Second Workshop on Computational Research in Linguistic Typology",
    "1761": "Search-Oriented Conversational AI (SCAI) 2",
    "2217": "The Fourth Workshop on Online Abuse and Harms (WOAH) a.k.a. ALW",
    "2487": "1st Workshop on Computational Approaches to Discourse",
    "2575": "Workshop on Insights from Negative Results in NLP",
    "2797": "Deep Learning Inside Out (DeeLIO): The First Workshop on Knowledge Extraction and Integration for Deep Learning Architectures",
    "2800": "Deep Learning Inside Out (DeeLIO): The First Workshop on Knowledge Extraction and Integration for Deep Learning Architectures",
    "2976": "BlackboxNLP 2020: Analyzing and interpreting neural networks for NLP",
    "3476": "Interactive and Executable Semantic Parsing (Int-Ex)",
    "3561": "BlackboxNLP 2020: Analyzing and interpreting neural networks for NLP",
}


@dataclass
class Session:
    name: str
    start: datetime
    end: datetime
    host: str


@dataclass
class Workshop:
    uid: str
    sessions: List[Session]
    description: str


@dataclass
class Paper:
    uid: str
    title: str
    authors: str
    abstract: str
    track: str
    kind: str
    link: str


def load_workshop_overview_excel() -> pd.DataFrame:
    wb = load_workbook(PATH_WORKSHOPS_OVERVIEW)
    ws = wb.worksheets[0]
    ws.delete_rows(1, 1)
    ws.delete_rows(27, 100)
    ws.delete_cols(7, 3)
    ws.delete_cols(8, 14)

    emnlp_workshops = pd.read_csv(PATH_WORKSHOPS_CSV)

    softconf_id_to_organizers = {
        row["softconfNumber"]: row["authors"] for _, row in emnlp_workshops.iterrows()
    }

    df = pd.DataFrame(
        ws.values,
        columns=[
            "Softconf Number",
            "UID",
            "Name",
            "Summary",
            "Authors",
            "URL",
            "Alias",
            "Old UID",
        ],
    )
    df = df.dropna(subset=["UID"])
    df["Softconf Number"] = df["Softconf Number"].fillna(-1)

    df["Softconf Number"] = df["Softconf Number"].apply(lambda x: int(x))
    df["Organizers"] = df["Softconf Number"].apply(
        lambda x: softconf_id_to_organizers[x]
    )

    return df


def build_workshops_basics() -> List[Dict[str, Any]]:
    workshops = load_workshop_overview_excel()
    schedule = load_schedule()
    zooms = get_zooms()

    data = []
    for _, row in workshops.iterrows():
        uid = row["UID"].strip()
        if uid == "WS-22":
            continue

        alias = row["Alias"]

        if alias is None:
            other = {"WS-4": "SCAI", "WS-1": "ConLL", "WS-13": "DeeLIO"}
            alias = other[row["UID"]]

        alias = alias.lower()
        sessions = [
            {
                "start_time": session.start,
                "end_time": session.end,
                "name": session.name,
                "hosts": session.host,
            }
            for session in schedule[uid].sessions
        ]

        entry = {
            "UID": uid,
            "title": row["Name"].strip(),
            "organizers": row["Organizers"].strip(),
            "abstract": row["Summary"],
            "website": row["URL"],
            "rocketchat_channel": f"workshop-{alias.lower()}",
            "alias": alias,
            "sessions": sessions,
        }

        if uid in zooms:
            entry["zoom_links"] = zooms[uid]

        data.append(entry)

    data.sort(key=lambda w: -int(w["UID"][2:]))

    return data


def load_schedule():
    wb = load_workbook(PATH_WORKSHOPS_SCHEDULE)

    data = {}
    for ws in wb.worksheets[4:]:
        workshop_id = ws["B2"].value
        assert workshop_id.startswith("WS-"), "Does not start with WS: " + workshop_id

        description = ws["B3"].value or ""
        ws.delete_rows(1, 6)
        ws.delete_cols(7, 100)
        df = pd.DataFrame(
            ws.values,
            columns=[
                "Session Name",
                "Day",
                "Start Time",
                "End Time",
                "Time Zone",
                "Host",
            ],
        )
        df.dropna(subset=["Session Name"], inplace=True)

        sessions = []
        for idx, row in df.iterrows():
            name = row["Session Name"].strip()
            host = row["Host"] or "TBD"

            day = row["Day"]
            start_time = row["Start Time"]
            end_time = row["End Time"]
            tz_name = row["Time Zone"]

            if not name or not tz_name:
                continue

            if isinstance(start_time, str):
                start_time = datetime.strptime(start_time, "%H:%M").time()
            if isinstance(end_time, str):
                end_time = datetime.strptime(end_time, "%H:%M").time()

            if isinstance(start_time, datetime):
                start_time = start_time.time()
            if isinstance(end_time, datetime):
                end_time = end_time.time()

            tz = pytz.timezone(tz_name)

            start = datetime.combine(day.date(), start_time)
            start = tz.localize(start)

            if start_time > end_time:
                day += timedelta(days=1)

            end = datetime.combine(day.date(), end_time)
            end = tz.localize(end)

            session = Session(name, start, end, host)
            sessions.append(session)

        workshop = Workshop(workshop_id, sessions, description)
        assert workshop_id not in data, (
            "workshop id already in data",
            workshop_id,
            data[workshop_id],
        )
        data[workshop_id] = workshop

    return data


def load_slideslive():
    # https://docs.google.com/spreadsheets/d/1Cp04DGRiDj8oY00-xDjTpjzCd_fjq3YhqOclhvFRK94/edit#gid=0
    df = pd.read_csv(PATH_SLIDESLIVE_WORKSHOPS)
    df = df.drop([0])

    workshop_df = load_workshop_overview_excel()

    ws_name_to_id = {
        row["Name"]: row["UID"].strip() for _, row in workshop_df.iterrows()
    }

    return df


def generate_workshop_papers(slideslive: pd.DataFrame):
    venues = []
    UIDs = []
    titles = []
    authors = []
    presentation_ids = []

    for _, row in slideslive.iterrows():
        if is_not_paper(row):
            continue

        title = row["Title"].replace("\n", " ")
        title = LatexNodes2Text().latex_to_text(title)
        title = fix_text(title).strip()
        author_list = [
            fix_text(e.strip()) for e in re.split(",| and | And ", row["Speakers"])
        ]

        ws = row["Organizer track name"].strip()
        uid = row["Unique ID"].strip()

        print(ws, uid)
        if ws == "WS-15" and str(uid) in fix.keys():
            continue

        venues.append(ws)
        UIDs.append(f"{ws}.{uid}")
        titles.append(title)
        authors.append("|".join(author_list))
        presentation_ids.append(
            row["SlidesLive link"].replace("https://slideslive.com/", "")
        )

    anthology_papers = get_anthology_workshop_papers()
    title_to_anthology_paper = {a.title.strip().lower(): a for a in anthology_papers}
    author_to_anthology_paper = {a.authors.lower(): a for a in anthology_papers}

    unmatched = []
    uid_to_anthology_paper = {}
    for uid, title, author in zip(UIDs, titles, authors):
        if uid.startswith(("WS-2")):
            continue

        if title.lower() in title_to_anthology_paper:
            assert uid not in uid_to_anthology_paper
            uid_to_anthology_paper[uid] = title_to_anthology_paper[title.lower()]
        else:
            unmatched.append((uid, title, author.lower()))

    for uid, title, author in list(unmatched):
        if author.lower() in author_to_anthology_paper:
            assert uid not in uid_to_anthology_paper, (
                uid,
                title,
                author,
                uid_to_anthology_paper[uid],
            )
            uid_to_anthology_paper[uid] = author_to_anthology_paper[author.lower()]
            unmatched.remove((uid, title, author.lower()))

    unmatched_df = pd.DataFrame(unmatched)
    unmatched_df.to_csv("unmatched_workshop_papers.csv", index=False)
    for e in unmatched:
        print(e)

    print(len(unmatched), len(uid_to_anthology_paper))

    abstracts = []
    urls = []
    for uid in UIDs:
        if uid in uid_to_anthology_paper:
            paper = uid_to_anthology_paper[uid]
            abstracts.append(paper.abstract)
            urls.append(paper.link)
        else:
            abstracts.append("")
            urls.append("")

    data = {
        "workshop": venues,
        "UID": UIDs,
        "title": titles,
        "authors": authors,
        "abstract": abstracts,
        "presentation_id": presentation_ids,
        "pdf_url": urls,
    }

    columns = [
        "workshop",
        "UID",
        "title",
        "authors",
        "abstract",
        "presentation_id",
        "pdf_url",
    ]
    df = pd.DataFrame(data, columns=columns)
    df = df.drop_duplicates(subset=["UID"])

    df.to_csv(PATH_YAMLS / "workshop_papers.csv", index=False)


def get_anthology_workshop_papers() -> List[Paper]:
    anthology = (
        Path(
            r"C:\Users\klie\AppData\Roaming\JetBrains\PyCharm2020.2\scratches\emnlp\acl-anthology"
        )
        / "data"
    )

    conference = "emnlp"
    year = 2020

    mapping = {
        "2020.conll-1": "WS-1",
        "2020.alw-1": "WS-17",
        "2020.blackboxnlp-1": "WS-25",
        "2020.clinicalnlp-1": "WS-12",
        "2020.cmcl-1": "WS-5",
        "2020.codi-1": "WS-16",
        "2020.deelio-1": "WS-13",
        "2020.eval4nlp-1": "WS-20",
        "2020.insights-1": "WS-3",
        "2020.intexsempar-1": "WS-6",
        "2020.louhi-1": "WS-19",
        "2020.nlpbt-1": "WS-23",
        "2020.nlpcovid19-1": "WS-26",
        "2020.nlpcss-1": "WS-18",
        "2020.nlposs-1": "WS-9",
        "2020.privatenlp-1": "WS-24",
        "2020.scai-1": "WS-4",
        "2020.sdp-1": "WS-7",
        "2020.sigtyp-1": "WS-11",
        "2020.splu-1": "WS-10",
        "2020.spnlp-1": "WS-21",
        "2020.sustainlp-1": "WS-15",
        "2020.wnut-1": "WS-14",
        "2020.findings-1": "findings",
    }

    papers = []
    for venue in mapping.keys():
        if venue.endswith("-1"):
            file_name = venue[:-2]
        else:
            file_name = venue

        path_to_xml = anthology / "xml" / f"{file_name}.xml"
        tree = ET.parse(path_to_xml)
        root = tree.getroot()
        collection_id = root.attrib["id"]

        for volume in root.findall("volume"):

            volume_id = volume.attrib["id"]

            for paper in volume.findall("paper"):
                paper_id = paper.attrib["id"]
                title = "".join(paper.find("title").itertext())
                uid = f"{collection_id}-{volume_id}.{paper_id}"
                authors = [
                    " ".join(author.itertext()) for author in paper.findall("author")
                ]
                authors = "|".join(authors)

                if paper.find("abstract") is not None:
                    abstract = "".join(paper.find("abstract").itertext())
                else:
                    abstract = ""

                link = f"https://www.aclweb.org/anthology/{uid}"

                track = mapping[venue]
                kind = None

                if track.startswith("W"):
                    kind = "workshop"
                elif track == "main":
                    kind = "long"
                else:
                    kind = "findings"

                assert kind

                paper = Paper(
                    uid=uid,
                    title=title,
                    authors=authors,
                    abstract=abstract,
                    track=track,
                    kind=kind,
                    link=link,
                )

                papers.append(paper)

    return papers


def is_not_paper(row) -> bool:
    uid = row["Unique ID"].lower()
    title = row["Title"].lower()

    return (
        ("invited" in uid)
        or ("challenge" in uid)
        or ("invited" in title)
        or ("keynote" in title)
        or ("keynote" in uid)
        or (row["Unique ID"] == "Shared task")
        or (title == "tba" and "paper" not in uid)
    )


def add_invited_talks(slideslive: pd.DataFrame):
    talks_per_workshop = defaultdict(list)

    for _, row in slideslive.iterrows():
        if not is_not_paper(row):
            continue

        title = row["Title"]
        speakers = row["Speakers"]
        presentation_id = row["SlidesLive link"].replace("https://slideslive.com/", "")

        talks_per_workshop[row["Organizer track name"]].append(
            {"title": title, "speakers": speakers, "presentation_id": presentation_id}
        )

    return talks_per_workshop


def get_zooms() -> Dict[str, List[str]]:
    df = pd.read_excel(PATH_ZOOM_ACCOUNTS_WITH_PASSWORDS, sheet_name="Workshops")

    zooms = defaultdict(list)
    for _, row in df.iterrows():
        uid = row["UID"].replace(".", "-").upper()
        zooms[uid].append(row["Personal Meeting LINK"])

        for i in range(row["# of accounts"] - 1):
            zooms[uid].append(row[f"Personal Meeting LINK.{i+1}"])

    return zooms


if __name__ == "__main__":
    # download_slideslive()
    # download_workshops()
    # download_zooms()

    # load_csv()
    data = build_workshops_basics()
    slideslive = load_slideslive()
    generate_workshop_papers(slideslive)
    talks = add_invited_talks(slideslive)

    fix_talks = slideslive[[is_not_paper(r) for _, r in slideslive.iterrows()]]
    fix_talks.to_csv(
        "yamls/fix_talks.csv",
        index=False,
        columns=["Organizer track name", "Unique ID", "Title", "Speakers"],
    )

    for ws in data:
        uid = ws["UID"]
        ws["prerecorded_talks"] = talks[uid]

    yaml.scalarstring.walk_tree(data)

    with open(PATH_YAMLS / "workshops.yml", "w") as f:
        yaml.dump(data, f, Dumper=ruamel.yaml.RoundTripDumper)
